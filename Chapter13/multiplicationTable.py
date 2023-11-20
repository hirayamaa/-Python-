import openpyxl
from openpyxl.styles import Font
import sys

if len(sys.argv) != 2:
    print('引数の数が間違っています')
    print('例：python multiplicationTable.py 6')
else:
    wb = openpyxl.Workbook()
    sheet = wb.active
    bold_font = Font(bold=True)
    num = int(sys.argv[1])
    # A列と1行目を太字にした見出しを作成する
    for row in range(2, num + 2):
        header_val = sheet.cell(row=row, column=1)
        header_val.font = bold_font
        header_val.value = row - 1
    for col in range(2, num + 2):
        header_val = sheet.cell(row=1, column=col)
        header_val.font = bold_font
        header_val.value = col - 1

    # 行ごとのループ
    for row in range(2, num + 2):
        # 列ごとのループ
        for col in range(2, num + 2):
            multi_val = sheet.cell(row=row, column=col)
            multi_val.value = (row - 1) * (col - 1)
    wb.save('Chapter13/multipleTable.xlsx')
    print('作成が完了しました')

