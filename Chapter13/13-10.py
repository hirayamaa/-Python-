import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide Column'
# 高さと幅を設定する
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')

wb = openpyxl.Workbook()
sheet = wb.active
# セルの結合
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells.'
wb.save('merged.xlsx')

# セルの結合解除
sheet.unmerge_cells('A1:D3')
sheet.unmerge_cells('C5:D5')
wb.save('merged.xlsx')

# ウィンドウ枠の固定
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
# A2より上の行を固定
sheet.freeze_panes = 'A2'
wb.save('freezeExample.xlsx')
