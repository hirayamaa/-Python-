import sys
import openpyxl

if len(sys.argv) != 4:
    print('Error')
else:
    blank_start_row = int(sys.argv[1])
    blank_row_num = int(sys.argv[2])
    read_wb = openpyxl.load_workbook('Chapter13/' + sys.argv[3])
    read_sheet = read_wb.active
    # 書き込み用のワークブック
    write_wb = openpyxl.Workbook()
    write_sheet = write_wb.active
    # 空白を挿入する前の行を書き込む
    for row in range(1, blank_start_row):
        for col in range(1, read_sheet.max_column + 1):
            write_sheet.cell(row=row, column=col).value \
                = read_sheet.cell(row=row, column=col).value
    # 挿入した空白以降の行を書き込む
    for row in range(blank_start_row, read_sheet.max_row + 1):
        for col in range(1, read_sheet.max_column + 1):
            write_sheet.cell(row=row + blank_row_num, column=col).value \
                = read_sheet.cell(row=row, column=col).value
    write_wb.save('Chapter13/insertedExcel.xlsx')
    print('作成完了しました')
