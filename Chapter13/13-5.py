import openpyxl

# 空白のワークブックを作成
wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb.active
print(sheet.title)
# シート名を変更
sheet.title = 'Spam Bacon Eggs Sheet'
print(wb.sheetnames)

print('###################')

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet.title = 'Spam Spam Spam'
# ワークブックを保存
wb.save('example_copy.xlsx')

wb = openpyxl.Workbook()
# 新しいシートの作成・削除
wb.create_sheet()
print(wb.sheetnames)
wb.create_sheet(index=0, title='First Sheet')
# ['First Sheet', 'Sheet', 'Sheet1']
print(wb.sheetnames)
wb.create_sheet(index=2, title='Middle Sheet')
# ['First Sheet', 'Sheet', 'Middle Sheet', 'Sheet1']
print(wb.sheetnames)
del wb['Middle Sheet']
print(wb.sheetnames)

print('###################')
# セルに値を書き込む
wb = openpyxl.Workbook()
# 書き込むシートの選択
sheet = wb['Sheet']
# セルの値を変更
sheet['A1'] = 'Hello World!'
print(sheet['A1'].value)
