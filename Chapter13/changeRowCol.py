import sys
import openpyxl

if len(sys.argv) != 2:
    print('引数の数が間違っています')
    print('例：python changeRowCol.py (ワークブックファイル名)')
else:
    # 引数からワークブックを読み込む
    read_wb = openpyxl.load_workbook('Chapter13/' + sys.argv[1])
    read_sheet = read_wb.active
    # 書き込む用のワークブックを作成する
    write_wb = openpyxl.Workbook()
    write_sheet = write_wb.active
    for row in range(1, read_sheet.max_row + 1):
        for col in range(1, read_sheet.max_column + 1):
            # 読み込んだx行y列のセルをy行x列のセルに書き込む
            write_sheet.cell(row=col, column=row).value \
                = read_sheet.cell(row=row, column=col).value

    write_wb.save('Chapter13/changedExcel.xlsx')
    print('作成完了しました')
