import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames)
sheet = wb['Sheet1']
print(sheet['A1'].value)
c = sheet['B1']
print(c.value)
# Row 1, Column 2 is Apples
print(f'Row {c.row}, Column {c.column} is {c.value}')
# Cell B1 is Apples
print(f'Cell {c.coordinate} is {c.value}')
print('################')

# cell()メソッドでループしてセルの値を取得する
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

print('################')
print(f'最大行数は{sheet.max_row}, 最大列数は{sheet.max_column}')

print('################')

print(get_column_letter(sheet.max_column))
print(column_index_from_string('A'))

print('################')

# シートから複数の行と列を取得する
sheet_tuple = tuple(sheet['A1':'C3'])
print(sheet_tuple)
for row_of_cell_objects in sheet['A1': 'C3']:
    for cell_obj in row_of_cell_objects:
        print(cell_obj.coordinate, cell_obj.value)
    print('--- END OF ROW ---')

print('################')
sheet_tuple = tuple(sheet.columns)[1]
print(sheet_tuple)
for cell_obj in sheet_tuple:
    print(cell_obj.value)

# openpyxlモジュールをインポートする。
# openpyxl.load_workbook()関数を呼び出す。
# Workbookオブジェクトを取得する。
# Workbookオブジェクトのプロパティactive、sheetnamesを用いる。
# Worksheetオブジェクトを取得する。
# インデックスやスライス、Worksheetオブジェクトのcell()メソッドにrowとcolumnのキーワード引数を渡す。
# Cellオブジェクトを取得する。
# Cellオブジェクトのvalue属性を読み出す。
