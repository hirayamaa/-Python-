import csv
import os
import openpyxl

os.makedirs('convertedCSV', exist_ok=True)

for excel_file in os.listdir('excelSpreadsheets'):
    # エクセルファイル以外は変換の対象外
    if not excel_file.endswith('.xlsx'):
        continue
    # ワークブックを読み込む
    wb = openpyxl.load_workbook(os.path.join('excelSpreadsheets', excel_file))
    # シートごとにループする
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # csvファイル名を作成する(エクセルファイル名_シート名)
        file_name = os.path.splitext(os.path.basename(excel_file))[0] + '_' + sheet_name + '.csv'
        csv_file = open(os.path.join('convertedCSV', file_name), 'w', newline='')
        csv_writer = csv.writer(csv_file)

        for row in range(1, sheet.max_row + 1):
            # CSVに書き込む１行をリストとして保持する
            excel_rows = []
            for col in range(1, sheet.max_column + 1):
                excel_rows.append(sheet.cell(row=row, column=col).value)
            csv_writer.writerow(excel_rows)
        csv_file.close()
